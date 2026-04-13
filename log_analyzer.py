"""
Log Analyzer - Identifica eventos de interés en logs de validadores VL550/CGI
Lee la lista de eventos desde un archivo .xlsx (lista-eventos-vl550.xlsx)

Uso:
    python log_analyzer.py <archivo_log.docx|.txt> [archivo_eventos.xlsx]

Requiere: pip install python-docx openpyxl
"""

import re
import sys
import os
import csv
import html as h
from collections import defaultdict
from docx import Document
from openpyxl import load_workbook

# Forzar UTF-8 en stdout (sys.stdout puede ser None en ejecutables sin consola)
if sys.stdout is not None and hasattr(sys.stdout, 'reconfigure') and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')


# ─────────────────────────────────────────────────────────────────────────────
# Lectura de eventos desde Excel
# Columnas esperadas: ID | Mensaje en log | Significado
# ─────────────────────────────────────────────────────────────────────────────
def load_events_from_xlsx(xlsx_path):
    """
    Lee el Excel y devuelve lista de dicts:
      { "id": int, "patron": str, "significado": str }
    Filas con 'Mensaje en log' vacío se omiten (no se puede buscar nada).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # saltar encabezado
        # columnas: ID, Mensaje en log, Significado
        if len(row) < 2:
            continue
        eid        = row[0]
        patron_raw = row[1]
        significado = row[2] if len(row) > 2 else ""

        # saltar filas sin ID o sin patrón de búsqueda
        if not eid or not patron_raw:
            continue

        events.append({
            "id":          int(eid),
            "patron":      str(patron_raw).strip(),
            "significado": str(significado).strip() if significado else "",
        })

    wb.close()
    return events


# ─────────────────────────────────────────────────────────────────────────────
# Lectura del log
# ─────────────────────────────────────────────────────────────────────────────
def extract_log_lines(file_path):
    """Extrae líneas de texto de un .docx o archivo de texto plano"""
    if file_path.lower().endswith(".docx"):
        doc = Document(file_path)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    else:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return [line.rstrip() for line in f if line.strip()]


# ─────────────────────────────────────────────────────────────────────────────
# Análisis
# ─────────────────────────────────────────────────────────────────────────────
TS_PATTERN = re.compile(r"^\[(\d{2}/\d{2}/\d{2}-\d{2}:\d{2}:\d{2}\.\d+)\]")


def analyze_log(log_lines, events):
    """
    Busca cada evento en las líneas del log.
    Retorna lista de dicts con todos los campos para el reporte.
    """
    # Compilar patrones una sola vez (búsqueda literal, case-insensitive)
    compiled = [
        (ev, re.compile(re.escape(ev["patron"]), re.IGNORECASE))
        for ev in events
    ]

    results = []
    for line_num, line in enumerate(log_lines, start=1):
        ts_match  = TS_PATTERN.match(line)
        timestamp = ts_match.group(1) if ts_match else "N/A"
        # mensaje limpio: quitar prefijo [ts][modulo][nivel]
        msg_clean = re.sub(r"^\[.*?\]\[.*?\]\[.*?\]", "", line).strip() or line

        for ev, pattern in compiled:
            if pattern.search(line):
                results.append({
                    "id":          ev["id"],
                    "patron":      ev["patron"],
                    "significado": ev["significado"],
                    "timestamp":   timestamp,
                    "linea_num":   line_num,
                    "mensaje":     msg_clean,
                    "linea":       line,
                })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Reporte en consola con tablas
# ─────────────────────────────────────────────────────────────────────────────
def _col_widths(rows, headers):
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)))
    return widths


def _print_table(headers, rows, title=""):
    widths = _col_widths(rows, headers)
    top = "┌" + "┬".join("─" * (w + 2) for w in widths) + "┐"
    mid = "├" + "┼".join("─" * (w + 2) for w in widths) + "┤"
    bot = "└" + "┴".join("─" * (w + 2) for w in widths) + "┘"

    def row_line(cells):
        return "│" + "│".join(f" {str(c):<{widths[i]}} " for i, c in enumerate(cells)) + "│"

    if title:
        print(f"\n  {title}")
    print(top)
    print(row_line(headers))
    print(mid)
    for r in rows:
        print(row_line(r))
    print(bot)


def print_report(results, log_lines):
    SEP = "═" * 100
    print(f"\n{SEP}")
    print(f"  ANÁLISIS DE LOG  │  {len(results)} ocurrencia(s) encontrada(s)  │  {len(log_lines)} líneas procesadas")
    print(SEP)

    if not results:
        print("  No se encontraron eventos de interés.")
        return

    # ── Resumen por ID ───────────────────────────────────────────────────────
    by_id = defaultdict(list)
    for r in results:
        by_id[r["id"]].append(r)

    summary_rows = []
    for eid in sorted(by_id.keys()):
        sample = by_id[eid][0]
        summary_rows.append([
            eid,
            len(by_id[eid]),
            sample["patron"][:50] + ("…" if len(sample["patron"]) > 50 else ""),
            sample["significado"][:60] + ("…" if len(sample["significado"]) > 60 else ""),
        ])

    _print_table(
        headers=["ID", "Ocurrencias", "Patrón buscado", "Significado"],
        rows=summary_rows,
        title="RESUMEN POR TIPO DE EVENTO"
    )

    # ── Detalle cronológico ──────────────────────────────────────────────────
    detail_rows = []
    for r in results:
        msg = r["mensaje"]
        detail_rows.append([
            r["id"],
            r["timestamp"],
            r["linea_num"],
            msg[:70] + ("…" if len(msg) > 70 else ""),
            r["significado"][:50] + ("…" if len(r["significado"]) > 50 else ""),
        ])

    _print_table(
        headers=["ID", "Timestamp", "Línea", "Mensaje", "Significado"],
        rows=detail_rows,
        title="DETALLE CRONOLÓGICO"
    )

    print(f"\n{SEP}")
    print(f"  Total ocurrencias: {len(results)}  │  Tipos distintos: {len(by_id)}")
    print(SEP)


# ─────────────────────────────────────────────────────────────────────────────
# Exportación CSV
# ─────────────────────────────────────────────────────────────────────────────
def export_csv(results, output_path="eventos_encontrados.csv"):
    fields = ["id", "patron", "significado", "timestamp", "linea_num", "mensaje", "linea"]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    print(f"\n  CSV exportado: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Exportación XLSX
# ─────────────────────────────────────────────────────────────────────────────
def export_xlsx(results, output_path="eventos_encontrados.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"

    headers = ["ID", "Patrón", "Significado", "Timestamp", "Línea Nro", "Mensaje"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")

    for col, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["id"])
        ws.cell(row=row_idx, column=2, value=r["patron"])
        ws.cell(row=row_idx, column=3, value=r["significado"])
        ws.cell(row=row_idx, column=4, value=r["timestamp"])
        ws.cell(row=row_idx, column=5, value=r["linea_num"])
        ws.cell(row=row_idx, column=6, value=r["mensaje"])

    # Ajustar ancho de columnas
    col_widths = [8, 40, 50, 22, 10, 80]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"  XLSX exportado: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Exportación HTML
# ─────────────────────────────────────────────────────────────────────────────
def export_html(results, log_file, output_path="reporte_eventos.html"):
    by_id = defaultdict(list)
    for r in results:
        by_id[r["id"]].append(r)

    # Paleta de colores cíclica para los badges de ID
    PALETTE = [
        "#1a6fa8","#2e7d32","#6a1b9a","#c62828","#00838f",
        "#ef6c00","#4527a0","#00695c","#ad1457","#37474f",
    ]
    id_color = {eid: PALETTE[i % len(PALETTE)] for i, eid in enumerate(sorted(by_id.keys()))}

    def badge(eid):
        c = id_color.get(eid, "#555")
        return f"<span class='badge' style='background:{c}'>{eid}</span>"

    # ── Resumen ──────────────────────────────────────────────────────────────
    def extract_value(patron, mensaje):
        """Extrae el valor que sigue al patrón en el mensaje."""
        try:
            idx = mensaje.lower().find(patron.lower())
            if idx == -1:
                return ""
            after = mensaje[idx + len(patron):].strip()
            # Tomar hasta el primer separador (coma, punto y coma, salto, corchete)
            val = re.split(r'[,;\n\[\]{}]', after)[0].strip()
            return val[:60]  # limitar longitud
        except Exception:
            return ""

    summary_rows = ""
    for eid in sorted(by_id.keys()):
        sample = by_id[eid][0]
        count  = len(by_id[eid])
        bar_w  = min(count * 18, 200)
        valor  = extract_value(sample['patron'], sample['mensaje'])
        summary_rows += (
            f"<tr>"
            f"<td class='ctr'>{badge(eid)}</td>"
            f"<td class='ctr'>"
            f"  <div class='bar-wrap'><div class='bar' style='width:{bar_w}px;background:{id_color[eid]}'></div>"
            f"  <span class='bar-num'>{count}</span></div>"
            f"</td>"
            f"<td class='mono'>{h.escape(sample['patron'])}</td>"
            f"<td class='mono val'>{h.escape(valor)}</td>"
            f"<td>{h.escape(sample['significado'])}</td>"
            f"</tr>\n"
        )

    # ── Detalle ──────────────────────────────────────────────────────────────
    detail_rows = ""
    for r in results:
        valor = extract_value(r['patron'], r['mensaje'])
        detail_rows += (
            f"<tr data-id='{r['id']}'>"
            f"<td class='ctr'>{badge(r['id'])}</td>"
            f"<td class='mono ts'>{h.escape(r['timestamp'])}</td>"
            f"<td class='ctr'>{r['linea_num']}</td>"
            f"<td class='mono msg'>{h.escape(r['mensaje'])}</td>"
            f"<td class='mono val'>{h.escape(valor)}</td>"
            f"<td class='sig'>{h.escape(r['significado'])}</td>"
            f"</tr>\n"
        )

    # ── Opciones de filtro por ID ─────────────────────────────────────────────
    filter_btns = "<button class='fbtn active' onclick=\"filterTable('all',this)\">Todos</button>\n"
    for eid in sorted(by_id.keys()):
        c = id_color[eid]
        filter_btns += (
            f"<button class='fbtn' style='--bc:{c}' onclick=\"filterTable({eid},this)\">"
            f"ID {eid} <span class='fcnt'>({len(by_id[eid])})</span></button>\n"
        )

    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reporte de Eventos — {h.escape(log_file)}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin:0; padding:0; }}
  body   {{ font-family: Segoe UI, Arial, sans-serif; background:#eef1f6; color:#222; padding:28px 32px; }}
  h1     {{ font-size:1.4rem; color:#1a3a5c; border-bottom:3px solid #1a6fa8;
            padding-bottom:10px; margin-bottom:6px; }}
  h2     {{ font-size:1rem; color:#1a3a5c; margin:32px 0 10px; text-transform:uppercase;
            letter-spacing:.05em; }}
  .meta  {{ background:#fff; border-left:4px solid #1a6fa8; padding:10px 18px;
            margin:14px 0 28px; border-radius:4px; font-size:.88rem; color:#444;
            display:flex; gap:24px; flex-wrap:wrap; }}
  .meta strong {{ color:#1a3a5c; }}

  /* ── Tarjetas de resumen rápido ── */
  .cards {{ display:flex; gap:16px; flex-wrap:wrap; margin-bottom:28px; }}
  .card  {{ background:#fff; border-radius:8px; padding:16px 22px;
            box-shadow:0 1px 4px rgba(0,0,0,.1); min-width:140px; text-align:center; }}
  .card .num  {{ font-size:2rem; font-weight:700; color:#1a6fa8; line-height:1; }}
  .card .lbl  {{ font-size:.78rem; color:#666; margin-top:4px; }}

  /* ── Tablas ── */
  .tbl-wrap {{ overflow-x:auto; border-radius:8px; box-shadow:0 1px 5px rgba(0,0,0,.1);
               margin-bottom:36px; }}
  table  {{ border-collapse:collapse; width:100%; background:#fff; font-size:.86rem; }}
  thead tr {{ background:#1a3a5c; }}
  th     {{ color:#fff; padding:10px 14px; text-align:left; white-space:nowrap;
            font-size:.82rem; font-weight:600; letter-spacing:.03em; }}
  td     {{ padding:8px 14px; border-bottom:1px solid #e4e8f0; vertical-align:top; }}
  tbody tr:last-child td {{ border-bottom:none; }}
  tbody tr:hover td {{ background:#f0f5ff; }}
  tbody tr.hidden   {{ display:none; }}

  /* ── Elementos ── */
  .badge {{ display:inline-block; padding:3px 9px; border-radius:12px; color:#fff;
            font-size:.78rem; font-weight:700; letter-spacing:.02em; }}
  .mono  {{ font-family: Consolas, 'Courier New', monospace; font-size:.82rem; }}
  .ts    {{ white-space:nowrap; color:#555; }}
  .msg   {{ word-break:break-word; max-width:420px; }}
  .sig   {{ color:#2e4a6e; font-size:.84rem; }}
  .val   {{ color:#555; font-size:.82rem; max-width:180px; word-break:break-word; }}
  .ctr   {{ text-align:center; }}

  /* ── Barra de ocurrencias ── */
  .bar-wrap {{ display:flex; align-items:center; gap:8px; }}
  .bar      {{ height:10px; border-radius:5px; min-width:4px; }}
  .bar-num  {{ font-weight:700; font-size:.85rem; color:#333; }}

  /* ── Filtros por ID ── */
  .filters {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:14px; }}
  .fbtn    {{ padding:5px 12px; border-radius:16px; border:2px solid var(--bc,#1a6fa8);
              background:#fff; color:var(--bc,#1a6fa8); font-size:.8rem; font-weight:600;
              cursor:pointer; transition:all .15s; }}
  .fbtn:hover, .fbtn.active {{ background:var(--bc,#1a6fa8); color:#fff; }}
  .fcnt  {{ font-weight:400; opacity:.85; }}

  /* ── Barra de búsqueda ── */
  .search-bar {{ display:flex; align-items:center; gap:10px; margin-bottom:10px;
                 background:#fff; border:1px solid #c8d0e0; border-radius:8px;
                 padding:6px 14px; box-shadow:0 1px 3px rgba(0,0,0,.07); }}
  .search-bar input {{ border:none; outline:none; font-size:.88rem; flex:1;
                       font-family:Consolas,'Courier New',monospace; color:#222; }}
  .search-bar label {{ font-size:.78rem; color:#666; white-space:nowrap; }}
  .search-bar select {{ border:1px solid #c8d0e0; border-radius:4px; font-size:.8rem;
                        padding:2px 6px; color:#333; background:#f7f9fc; cursor:pointer; }}
  .search-bar .s-count {{ font-size:.78rem; color:#888; white-space:nowrap; }}
  .search-bar button {{ border:none; background:none; cursor:pointer; color:#888;
                        font-size:1rem; padding:0 4px; }}
  .search-bar button:hover {{ color:#c62828; }}
  mark {{ background:#fff176; border-radius:2px; padding:0 1px; }}
</style>
</head>
<body>

<h1>Reporte de Eventos de Log</h1>
<div class="meta">
  <span><strong>Archivo:</strong> {h.escape(log_file)}</span>
  <span><strong>Ocurrencias totales:</strong> {len(results)}</span>
  <span><strong>Tipos de evento:</strong> {len(by_id)}</span>
</div>

<div class="cards">
  <div class="card"><div class="num">{len(results)}</div><div class="lbl">Ocurrencias totales</div></div>
  <div class="card"><div class="num">{len(by_id)}</div><div class="lbl">Tipos de evento</div></div>
  <div class="card"><div class="num">{max((len(v) for v in by_id.values()), default=0)}</div><div class="lbl">Máx. ocurrencias (un tipo)</div></div>
</div>

<h2>Resumen por Tipo de Evento</h2>
<div class="tbl-wrap">
<table>
  <thead><tr><th>ID</th><th>Ocurrencias</th><th>Patrón buscado</th><th>Valor</th><th>Significado</th></tr></thead>
  <tbody>{summary_rows}</tbody>
</table>
</div>

<h2>Detalle Cronológico</h2>
<div class="filters" id="filters">{filter_btns}</div>
<div class="search-bar">
  <span>🔍</span>
  <input type="text" id="search-input" placeholder="Buscar en Timestamp, Línea, Mensaje o Valor…" oninput="applySearch()">
  <label for="search-col">en:</label>
  <select id="search-col" onchange="applySearch()">
    <option value="all">Todos los campos</option>
    <option value="1">Timestamp</option>
    <option value="2">Línea</option>
    <option value="3">Mensaje</option>
    <option value="4">Valor</option>
  </select>
  <span class="s-count" id="search-count"></span>
  <button onclick="clearSearch()" title="Limpiar búsqueda">✕</button>
</div>
<div class="tbl-wrap">
<table id="detail-table">
  <thead><tr><th>ID</th><th>Timestamp</th><th>Línea</th><th>Mensaje</th><th>Valor</th><th>Significado</th></tr></thead>
  <tbody>{detail_rows}</tbody>
</table>
</div>

<script>
var _activeId = 'all';

function filterTable(id, btn) {{
  document.querySelectorAll('.fbtn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  _activeId = id;
  applySearch();
}}

function applySearch() {{
  var term = document.getElementById('search-input').value.trim().toLowerCase();
  var col  = document.getElementById('search-col').value;
  var rows = document.querySelectorAll('#detail-table tbody tr');
  var visible = 0;

  rows.forEach(function(tr) {{
    // Filtro por ID
    var idMatch = (_activeId === 'all' || tr.dataset.id == _activeId);

    // Filtro por búsqueda
    var searchMatch = true;
    if (term) {{
      var cells = tr.querySelectorAll('td');
      // col indices: 1=Timestamp, 2=Línea, 3=Mensaje, 4=Valor
      var targets = col === 'all' ? [1,2,3,4] : [parseInt(col)];
      searchMatch = targets.some(function(i) {{
        return cells[i] && cells[i].textContent.toLowerCase().includes(term);
      }});
    }}

    var show = idMatch && searchMatch;
    tr.classList.toggle('hidden', !show);
    if (show) visible++;

    // Highlight
    [1,2,3,4].forEach(function(i) {{
      if (!tr.classList.contains('hidden') && term && (col === 'all' || parseInt(col) === i)) {{
        var cell = tr.querySelectorAll('td')[i];
        if (cell) cell.innerHTML = highlight(cell.textContent, term);
      }} else if (tr.querySelectorAll('td')[i]) {{
        var cell = tr.querySelectorAll('td')[i];
        if (cell.querySelector('mark')) cell.textContent = cell.textContent;
      }}
    }});
  }});

  var total = Array.from(rows).filter(r => !r.classList.contains('hidden')).length;
  document.getElementById('search-count').textContent = term ? total + ' resultado(s)' : '';
}}

function highlight(text, term) {{
  if (!term) return escHtml(text);
  var re = new RegExp('(' + term.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&') + ')', 'gi');
  return escHtml(text).replace(re, '<mark>$1</mark>');
}}

function escHtml(s) {{
  return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function clearSearch() {{
  document.getElementById('search-input').value = '';
  document.getElementById('search-count').textContent = '';
  applySearch();
}}
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"  HTML exportado: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":

    # ── Archivo de log ───────────────────────────────────────────────────────
    if len(sys.argv) > 1:
        log_file = sys.argv[1]
    else:
        candidates = [f for f in os.listdir(".")
                      if f.lower().endswith((".docx", ".txt"))
                      and "lista" not in f.lower()
                      and "evento" not in f.lower()]
        if not candidates:
            print("Uso: python log_analyzer.py <log.docx|.txt> [eventos.xlsx]")
            sys.exit(1)
        log_file = candidates[0]
        print(f"Log detectado automáticamente: {log_file}")

    if not os.path.exists(log_file):
        print(f"Error: no se encontró '{log_file}'")
        sys.exit(1)

    # ── Archivo de eventos Excel ─────────────────────────────────────────────
    if len(sys.argv) > 2:
        xlsx_file = sys.argv[2]
    else:
        # buscar automáticamente cualquier .xlsx en el directorio
        xlsx_candidates = [f for f in os.listdir(".") if f.lower().endswith(".xlsx")]
        if not xlsx_candidates:
            print("Error: no se encontró ningún archivo .xlsx con la lista de eventos.")
            print("Uso: python log_analyzer.py <log.docx> <eventos.xlsx>")
            sys.exit(1)
        xlsx_file = xlsx_candidates[0]
        print(f"Excel de eventos detectado automáticamente: {xlsx_file}")

    if not os.path.exists(xlsx_file):
        print(f"Error: no se encontró '{xlsx_file}'")
        sys.exit(1)

    # ── Ejecutar ─────────────────────────────────────────────────────────────
    print(f"\nCargando eventos desde: {xlsx_file}")
    events = load_events_from_xlsx(xlsx_file)
    print(f"Eventos cargados: {len(events)}")

    print(f"Leyendo log: {log_file}")
    log_lines = extract_log_lines(log_file)
    print(f"Líneas extraídas: {len(log_lines)}")

    # ── Archivo de salida ───────────────────────────────────────────────────
    if len(sys.argv) > 3:
        output_file = sys.argv[3]
    else:
        output_file = "reporte_eventos.html"

    results = analyze_log(log_lines, events)
    print_report(results, log_lines)

    if output_file.lower().endswith(".xlsx"):
        export_xlsx(results, output_file)
    else:
        export_html(results, log_file, output_file)
