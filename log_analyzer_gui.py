"""
GUI para log_analyzer.py — soporta múltiples logs en un solo reporte
Uso: python log_analyzer_gui.py
"""

import re
import sys
import os
import csv
import glob
import html as h
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import defaultdict
from docx import Document
from openpyxl import load_workbook

# Forzar UTF-8 en stdout (puede ser None en ejecutables sin consola)
if sys.stdout is not None and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


# ─────────────────────────────────────────────────────────────────────────────
# Funciones del analizador (incrustadas para compatibilidad con PyInstaller)
# ─────────────────────────────────────────────────────────────────────────────

def load_events_from_xlsx(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        eid = row[0]
        patron_raw = row[1]
        significado = row[2] if len(row) > 2 else ""
        if not eid:
            continue
        # Si la columna B está vacía, usar el significado (col C) como patrón
        if not patron_raw:
            if not significado:
                continue
            patron_raw = significado
        events.append({
            "id":          int(eid),
            "patron":      str(patron_raw).strip(),
            "significado": str(significado).strip() if significado else "",
        })
    wb.close()
    return events


def extract_log_lines(file_path):
    if file_path.lower().endswith(".docx"):
        doc = Document(file_path)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    else:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return [line.rstrip() for line in f if line.strip()]


_TS_PATTERN = re.compile(r"^\[(\d{2}/\d{2}/\d{2}-\d{2}:\d{2}:\d{2}\.\d+)\]")


def analyze_log(log_lines, events):
    import unicodedata

    def _normalize(s):
        return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii").lower()

    def _patron_to_regex(patron):
        # Escapar todo, luego restaurar * y ? como wildcards
        # * = cualquier cantidad de caracteres, ? = un solo carácter
        parts = re.split(r'(\*|\?)', patron)
        result = ""
        for part in parts:
            if part == '*':
                result += '.*'
            elif part == '?':
                result += '.'
            else:
                result += re.escape(part)
        return result

    compiled = [
        (ev, re.compile(_patron_to_regex(_normalize(ev["patron"]))))
        for ev in events
    ]
    results = []
    for line_num, line in enumerate(log_lines, start=1):
        ts_match  = _TS_PATTERN.match(line)
        timestamp = ts_match.group(1) if ts_match else "N/A"
        msg_clean = re.sub(r"^\[.*?\]\[.*?\]\[.*?\]", "", line).strip() or line
        line_norm = _normalize(line)
        for ev, pattern in compiled:
            if pattern.search(line_norm):
                results.append({
                    "id":          ev["id"],
                    "patron":      ev["patron"],
                    "significado": ev["significado"],
                    "timestamp":   timestamp,
                    "linea_num":   line_num,
                    "mensaje":     msg_clean,
                    "linea":       line,
                })
    return results


def export_xlsx(results, output_path="eventos_encontrados.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    headers = ["ID", "Patrón", "Significado", "Timestamp", "Línea Nro", "Mensaje"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    for col, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["id"])
        ws.cell(row=row_idx, column=2, value=r["patron"])
        ws.cell(row=row_idx, column=3, value=r["significado"])
        ws.cell(row=row_idx, column=4, value=r["timestamp"])
        ws.cell(row=row_idx, column=5, value=r["linea_num"])
        ws.cell(row=row_idx, column=6, value=r["mensaje"])
    col_widths = [8, 40, 50, 22, 10, 80]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)


def export_html(results, log_file, output_path="reporte_eventos.html"):
    by_id = defaultdict(list)
    for r in results:
        by_id[r["id"]].append(r)

    PALETTE = [
        "#1a6fa8","#2e7d32","#6a1b9a","#c62828","#00838f",
        "#ef6c00","#4527a0","#00695c","#ad1457","#37474f",
    ]
    id_color = {eid: PALETTE[i % len(PALETTE)] for i, eid in enumerate(sorted(by_id.keys()))}

    def badge(eid):
        c = id_color.get(eid, "#555")
        return f"<span class='badge' style='background:{c}'>{eid}</span>"

    _VALUE_PATTERNS = {
        3:  r'Tarjeta Mifare detectada\.\s*UID=(\S+)',
        7:  r'COMPANY:\s*(\S+)',
        10: r'ulLastFour:\s*(\S+)',
        11: r'"serial_number"\s*:\s*"([^"]+)"',
        12: r'appVersion\s*=\s*(v[\d\.\-k]+)',
        14: r'QR Record serialNumber:\s*(\S+)',
        20: r'FARE:\s*(\S+)',
        22: r't\.counter:(\S+)',
        23: r'CONTADOR_BOLETOS,\s*Value:\s*(\S+)',
        26: r'merchantName:\s*(.+)',
        27: r'driver\s*=\s*(\S+)',
        28: r'Name:\s*EVENTS_NUMBER,\s*Value:\s*(\S+)',
        29: r'"versionFW"\s*:\s*"(v[^"]+)"',
        34: r'Name:\s*SERVICE_ID,\s*Value:\s*(\S+)',
        40: r'Estado del validador:\s*(\d+)',
        43: r'Tabla:RL\s+id:9\s+currVersion:(\S+)',
        44: r'Tabla:AL\s+id:11\s+currVersion:(\S+)',
        45: r'Tabla:CO\s+id:3\s+currVersion:(\S+)',
        46: r'Tabla:CD\s+id:15\s+currVersion:(\S+)',
        47: r'Tabla:LR\s+id:23\s+currVersion:(\S+)',
        48: r'Tabla:RS\s+id:16\s+currVersion:(\S+)',
        49: r'Tabla:GP\s+id:1\s+currVersion:(\S+)',
        50: r'Tabla:SG\s+id:20\s+currVersion:(\S+)',
        51: r'Tabla:LI\s+id:18\s+currVersion:(\S+)',
        52: r'Tabla:OL\s+id:10\s+currVersion:(\S+)',
        58: r'\\?"latitude\\?"\s*:\s*(-?[0-9]+\.[0-9]+)',
        59: r'\\?"longitude\\?"\s*:\s*(-?[0-9]+\.[0-9]+)',
    }
    _compiled_value = {eid: re.compile(pat, re.IGNORECASE) for eid, pat in _VALUE_PATTERNS.items()}

    def extract_value(eid, mensaje):
        pat = _compiled_value.get(eid)
        if pat is None:
            return ""
        try:
            m = pat.search(mensaje)
            return m.group(1).strip()[:80] if m else ""
        except Exception:
            return ""

    summary_rows = ""
    for eid in sorted(by_id.keys()):
        sample = by_id[eid][0]
        count  = len(by_id[eid])
        bar_w  = min(count * 18, 200)
        valor  = extract_value(eid, sample['linea'])
        summary_rows += (
            f"<tr>"
            f"<td class='ctr'>{badge(eid)}</td>"
            f"<td class='ctr'><div class='bar-wrap'>"
            f"<div class='bar' style='width:{bar_w}px;background:{id_color[eid]}'></div>"
            f"<span class='bar-num'>{count}</span></div></td>"
            f"<td class='mono'>{h.escape(sample['patron'])}</td>"
            f"<td class='mono val'>{h.escape(valor)}</td>"
            f"<td>{h.escape(sample['significado'])}</td>"
            f"</tr>\n"
        )

    multi_source = any(r.get("_source") for r in results)
    detail_rows = ""
    for r in results:
        valor = extract_value(r['id'], r['linea'])
        source_td = f"<td class='src'>{h.escape(r.get('_source',''))}</td>" if multi_source else ""
        src_data = f"data-src='{h.escape(r.get('_source',''))}'" if multi_source else ""
        detail_rows += (
            f"<tr data-id='{r['id']}' data-ts='{h.escape(r['timestamp'])}' "
            f"data-ln='{r['linea_num']}' data-msg='{h.escape(r['mensaje'][:80])}' "
            f"data-val='{h.escape(valor)}' {src_data}>"
            f"<td class='ctr'>{badge(r['id'])}</td>"
            f"{source_td}"
            f"<td class='mono ts'>{h.escape(r['timestamp'])}</td>"
            f"<td class='ctr'>{r['linea_num']}</td>"
            f"<td class='mono msg'>{h.escape(r['mensaje'])}</td>"
            f"<td class='mono val'>{h.escape(valor)}</td>"
            f"<td class='sig'>{h.escape(r['significado'])}</td>"
            f"</tr>\n"
        )
    source_th = "<th class='sortable' data-col='src' onclick='sortTable(this)'>Fuente<span class='sort-icon'></span></th>" if multi_source else ""

    filter_checks = ""
    for eid in sorted(by_id.keys()):
        c = id_color[eid]
        sig = h.escape(by_id[eid][0]['significado'])
        cnt = len(by_id[eid])
        filter_checks += (
            f"<label class='chk-label' style='--bc:{c}'>"
            f"<input type='checkbox' class='ev-chk' value='{eid}' checked onchange='applySearch()'>"
            f"<span class='chk-badge' style='background:{c}'>{eid}</span>"
            f"<span class='chk-sig'>{sig}</span>"
            f"<span class='chk-cnt'>({cnt})</span>"
            f"</label>\n"
        )

    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reporte de Eventos — {h.escape(log_file)}</title>
<style>
  *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Segoe UI,Arial,sans-serif;background:#eef1f6;color:#222;padding:28px 32px}}
  h1{{font-size:1.4rem;color:#1a3a5c;border-bottom:3px solid #1a6fa8;padding-bottom:10px;margin-bottom:6px}}
  h2{{font-size:1rem;color:#1a3a5c;margin:32px 0 10px;text-transform:uppercase;letter-spacing:.05em}}
  .meta{{background:#fff;border-left:4px solid #1a6fa8;padding:10px 18px;margin:14px 0 28px;border-radius:4px;font-size:.88rem;color:#444;display:flex;gap:24px;flex-wrap:wrap}}
  .meta strong{{color:#1a3a5c}}
  .cards{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:28px}}
  .card{{background:#fff;border-radius:8px;padding:16px 22px;box-shadow:0 1px 4px rgba(0,0,0,.1);min-width:140px;text-align:center}}
  .card .num{{font-size:2rem;font-weight:700;color:#1a6fa8;line-height:1}}
  .card .lbl{{font-size:.78rem;color:#666;margin-top:4px}}
  .tbl-wrap{{overflow-x:auto;border-radius:8px;box-shadow:0 1px 5px rgba(0,0,0,.1);margin-bottom:36px}}
  table{{border-collapse:collapse;width:100%;background:#fff;font-size:.86rem}}
  thead tr{{background:#1a3a5c}}
  th{{color:#fff;padding:10px 14px;text-align:left;white-space:nowrap;font-size:.82rem;font-weight:600;letter-spacing:.03em}}
  td{{padding:8px 14px;border-bottom:1px solid #e4e8f0;vertical-align:top}}
  tbody tr:last-child td{{border-bottom:none}}
  tbody tr:hover td{{background:#f0f5ff}}
  tbody tr.hidden{{display:none}}
  .badge{{display:inline-block;padding:3px 9px;border-radius:12px;color:#fff;font-size:.78rem;font-weight:700;letter-spacing:.02em}}
  .mono{{font-family:Consolas,'Courier New',monospace;font-size:.82rem}}
  .ts{{white-space:nowrap;color:#555}}
  .msg{{word-break:break-word;max-width:420px}}
  .sig{{color:#2e4a6e;font-size:.84rem}}
  .val{{color:#555;font-size:.82rem;max-width:180px;word-break:break-word}}
  .src{{font-size:.78rem;color:#888;white-space:nowrap;font-family:Consolas,'Courier New',monospace}}
  .ctr{{text-align:center}}
  .bar-wrap{{display:flex;align-items:center;gap:8px}}
  .bar{{height:10px;border-radius:5px;min-width:4px}}
  .bar-num{{font-weight:700;font-size:.85rem;color:#333}}
  .filter-panel{{background:#fff;border:1px solid #c8d0e0;border-radius:8px;padding:12px 16px;margin-bottom:14px;box-shadow:0 1px 3px rgba(0,0,0,.07)}}
  .fp-header{{display:flex;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap}}
  .fp-title{{font-size:.82rem;font-weight:600;color:#1a3a5c;text-transform:uppercase;letter-spacing:.04em}}
  .fp-actions{{display:flex;gap:6px}}
  .fp-actions button{{padding:3px 10px;border-radius:12px;border:1px solid #c8d0e0;background:#f7f9fc;color:#444;font-size:.78rem;cursor:pointer}}
  .fp-actions button:hover{{background:#e0e8f5}}
  .chk-list{{display:flex;flex-wrap:wrap;gap:6px}}
  .chk-label{{display:flex;align-items:center;gap:5px;cursor:pointer;background:#f7f9fc;border:1.5px solid var(--bc,#1a6fa8);border-radius:20px;padding:4px 10px 4px 6px;font-size:.8rem;transition:background .15s;user-select:none}}
  .chk-label:hover{{background:#e8f0fb}}
  .chk-label input{{accent-color:var(--bc,#1a6fa8);width:14px;height:14px;cursor:pointer}}
  .chk-badge{{display:inline-block;padding:1px 7px;border-radius:10px;color:#fff;font-size:.75rem;font-weight:700}}
  .chk-sig{{color:#333;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
  .chk-cnt{{color:#888;font-size:.75rem}}
  .search-bar{{display:flex;align-items:center;gap:10px;margin-bottom:10px;background:#fff;border:1px solid #c8d0e0;border-radius:8px;padding:6px 14px;box-shadow:0 1px 3px rgba(0,0,0,.07)}}
  .search-bar input{{border:none;outline:none;font-size:.88rem;flex:1;font-family:Consolas,'Courier New',monospace;color:#222}}
  .search-bar label{{font-size:.78rem;color:#666;white-space:nowrap}}
  .search-bar select{{border:1px solid #c8d0e0;border-radius:4px;font-size:.8rem;padding:2px 6px;color:#333;background:#f7f9fc;cursor:pointer}}
  .search-bar .s-count{{font-size:.78rem;color:#888;white-space:nowrap}}
  .search-bar button{{border:none;background:none;cursor:pointer;color:#888;font-size:1rem;padding:0 4px}}
  .search-bar button:hover{{color:#c62828}}
  mark{{background:#fff176;border-radius:2px;padding:0 1px}}
  th.sortable{{cursor:pointer;user-select:none;white-space:nowrap}}
  th.sortable:hover{{background:#243f5e}}
  th.sortable .sort-icon{{display:inline-block;margin-left:5px;font-size:.75rem;opacity:.45;vertical-align:middle}}
  th.sortable.asc .sort-icon::after{{content:'▲';opacity:1}}
  th.sortable.desc .sort-icon::after{{content:'▼';opacity:1}}
  th.sortable:not(.asc):not(.desc) .sort-icon::after{{content:'⇅'}}
</style>
</head>
<body>
<h1>Reporte de Eventos de Log</h1>
<div class="meta">
  <span><strong>Archivo:</strong> {h.escape(log_file)}</span>
  <span><strong>Ocurrencias totales:</strong> {len(results)}</span>
  <span><strong>Tipos de evento:</strong> {len(by_id)}</span>
</div>
<div class="cards">
  <div class="card"><div class="num">{len(results)}</div><div class="lbl">Ocurrencias totales</div></div>
  <div class="card"><div class="num">{len(by_id)}</div><div class="lbl">Tipos de evento</div></div>
  <div class="card"><div class="num">{max((len(v) for v in by_id.values()), default=0)}</div><div class="lbl">Máx. ocurrencias (un tipo)</div></div>
</div>
<h2>Resumen por Tipo de Evento</h2>
<div class="tbl-wrap">
<table>
  <thead><tr><th>ID</th><th>Ocurrencias</th><th>Patrón buscado</th><th>1er Valor Registrado</th><th>Significado</th></tr></thead>
  <tbody>{summary_rows}</tbody>
</table>
</div>
<h2>Detalle Cronológico</h2>
<div class="filter-panel">
  <div class="fp-header">
    <span class="fp-title">Filtrar por tipo de evento</span>
    <div class="fp-actions">
      <button onclick="selectAll(true)">Seleccionar todos</button>
      <button onclick="selectAll(false)">Deseleccionar todos</button>
    </div>
  </div>
  <div class="chk-list">{filter_checks}</div>
</div>
<div class="search-bar">
  <span>🔍</span>
  <input type="text" id="search-input" placeholder="Buscar en Timestamp, Línea, Mensaje o Valor…" oninput="applySearch()">
  <label for="search-col">en:</label>
  <select id="search-col" onchange="applySearch()">
    <option value="all">Todos los campos</option>
    <option value="1">Timestamp</option>
    <option value="2">Línea</option>
    <option value="3">Mensaje</option>
    <option value="4">Valor</option>
  </select>
  <span class="s-count" id="search-count"></span>
  <button onclick="clearSearch()" title="Limpiar búsqueda">✕</button>
</div>
<div class="tbl-wrap">
<table id="detail-table">
  <thead><tr><th>ID</th>{source_th}<th class='sortable' data-col='ts' onclick='sortTable(this)'>Timestamp<span class='sort-icon'></span></th><th class='sortable' data-col='ln' onclick='sortTable(this)'>Línea<span class='sort-icon'></span></th><th class='sortable' data-col='msg' onclick='sortTable(this)'>Mensaje<span class='sort-icon'></span></th><th class='sortable' data-col='val' onclick='sortTable(this)'>Valor<span class='sort-icon'></span></th><th>Significado</th></tr></thead>
  <tbody>{detail_rows}</tbody>
</table>
</div>
<script>
function getCheckedIds(){{return Array.from(document.querySelectorAll('.ev-chk:checked')).map(function(c){{return c.value;}});}}
function selectAll(state){{document.querySelectorAll('.ev-chk').forEach(function(c){{c.checked=state;}});applySearch();}}
function applySearch(){{
  var term=document.getElementById('search-input').value.trim().toLowerCase();
  var col=document.getElementById('search-col').value;
  var rows=document.querySelectorAll('#detail-table tbody tr');
  var checked=getCheckedIds();
  rows.forEach(function(tr){{
    var idMatch=checked.includes(tr.dataset.id);
    var searchMatch=true;
    if(term){{
      var cells=tr.querySelectorAll('td');
      var targets=col==='all'?[1,2,3,4]:[parseInt(col)];
      searchMatch=targets.some(function(i){{return cells[i]&&cells[i].textContent.toLowerCase().includes(term);}});
    }}
    var show=idMatch&&searchMatch;
    tr.classList.toggle('hidden',!show);
    [1,2,3,4].forEach(function(i){{
      if(!tr.classList.contains('hidden')&&term&&(col==='all'||parseInt(col)===i)){{
        var cell=tr.querySelectorAll('td')[i];
        if(cell)cell.innerHTML=highlight(cell.textContent,term);
      }}else if(tr.querySelectorAll('td')[i]){{
        var cell=tr.querySelectorAll('td')[i];
        if(cell.querySelector('mark'))cell.textContent=cell.textContent;
      }}
    }});
  }});
  var total=Array.from(rows).filter(function(r){{return!r.classList.contains('hidden');}}).length;
  document.getElementById('search-count').textContent=term?total+' resultado(s)':'';
}}
function highlight(text,term){{
  if(!term)return escHtml(text);
  var escaped=term.replace(/[.*+?^${{}}()|[\]\\\\]/g,'\\\\$&');
  return escHtml(text).replace(new RegExp('('+escaped+')','gi'),'<mark>$1</mark>');
}}
function escHtml(s){{return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}
function clearSearch(){{document.getElementById('search-input').value='';document.getElementById('search-count').textContent='';applySearch();}}
function sortTable(th){{
  var col=th.dataset.col;
  var asc=!th.classList.contains('asc');
  document.querySelectorAll('#detail-table th.sortable').forEach(function(h){{h.classList.remove('asc','desc');}});
  th.classList.add(asc?'asc':'desc');
  var tbody=document.querySelector('#detail-table tbody');
  var rows=Array.from(tbody.querySelectorAll('tr'));
  rows.sort(function(a,b){{
    var av=a.dataset[col]||'', bv=b.dataset[col]||'';
    if(col==='ln'){{return asc?(parseInt(av)||0)-(parseInt(bv)||0):(parseInt(bv)||0)-(parseInt(av)||0);}}
    return asc?av.localeCompare(bv):bv.localeCompare(av);
  }});
  rows.forEach(function(r){{tbody.appendChild(r);}});
}}
</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)


def _show_done_dialog(parent, output_file, n_logs, n_results, n_lines):
    """Ventana de éxito con botones Abrir reporte / Abrir carpeta."""
    import subprocess

    dlg = tk.Toplevel(parent)
    dlg.title("Análisis completado")
    dlg.resizable(False, False)
    dlg.grab_set()

    frm = ttk.Frame(dlg, padding=20)
    frm.pack()

    # Ícono: pulgar arriba sobre fondo verde
    icon_frame = tk.Frame(frm, bg="#27ae60", width=48, height=48)
    icon_frame.pack_propagate(False)
    icon_frame.pack(pady=(0, 10))
    tk.Label(icon_frame, text="👍", font=("Segoe UI Emoji", 20),
             bg="#27ae60", fg="white").place(relx=0.5, rely=0.5, anchor="center")

    # Mensaje
    msg = (
        f"Logs analizados: {n_logs}\n"
        f"Ocurrencias encontradas: {n_results}\n"
        f"Líneas procesadas: {n_lines}\n\n"
        f"Reporte guardado en:\n{output_file}"
    )
    ttk.Label(frm, text=msg, justify="center",
              font=("Segoe UI", 9)).pack(pady=(0, 14))

    # Botones
    btn_frame = ttk.Frame(frm)
    btn_frame.pack()

    def open_report():
        try:
            os.startfile(output_file)
        except Exception:
            pass

    def open_folder():
        folder = os.path.dirname(os.path.abspath(output_file))
        try:
            subprocess.Popen(f'explorer /select,"{os.path.abspath(output_file)}"')
        except Exception:
            os.startfile(folder)

    ttk.Button(btn_frame, text="📄 Abrir reporte", width=18,
               command=open_report).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="📂 Abrir carpeta", width=18,
               command=open_folder).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Cerrar", width=10,
               command=dlg.destroy).pack(side="left", padx=4)

    # Centrar sobre la ventana padre
    dlg.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width() - dlg.winfo_width()) // 2
    y = parent.winfo_y() + (parent.winfo_height() - dlg.winfo_height()) // 2
    dlg.geometry(f"+{x}+{y}")

    # Sonido alegre
    try:
        import winsound
        winsound.PlaySound("SystemExclamation", winsound.SND_ALIAS | winsound.SND_ASYNC)
    except Exception:
        pass


def run_analysis(log_files, xlsx_file, output_file, status_var, btn_run, root,
                 api_state=None):
    """Ejecuta el análisis sobre múltiples logs y genera un único reporte."""
    import tempfile, shutil
    tmp_xlsx = None
    try:
        # Copiar xlsx a temporal para evitar errno 13 si está abierto en Excel
        tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_xlsx.close()
        shutil.copy2(xlsx_file, tmp_xlsx.name)

        status_var.set("Cargando eventos...")
        root.update_idletasks()
        events = load_events_from_xlsx(tmp_xlsx.name)

        all_results   = []
        total_lines   = 0
        log_label     = ", ".join(os.path.basename(f) for f in log_files)

        for i, lf in enumerate(log_files, 1):
            status_var.set(f"Leyendo log {i}/{len(log_files)}: {os.path.basename(lf)}…")
            root.update_idletasks()
            lines = extract_log_lines(lf)
            total_lines += len(lines)

            status_var.set(f"Analizando {i}/{len(log_files)}: {os.path.basename(lf)}…")
            root.update_idletasks()
            results = analyze_log(lines, events)

            # Anotar de qué archivo proviene cada resultado
            for r in results:
                r["_source"] = os.path.basename(lf)
            all_results.extend(results)

        status_var.set("Exportando reporte...")
        root.update_idletasks()

        if output_file.lower().endswith(".xlsx"):
            export_xlsx(all_results, output_file)
        else:
            export_html(all_results, log_label, output_file)

        status_var.set(
            f"Listo. {len(all_results)} ocurrencias en {total_lines} líneas "
            f"({len(log_files)} log(s))."
        )
        _show_done_dialog(root, output_file, len(log_files), len(all_results), total_lines)
        # Acumular métricas para el reporte API
        if api_state and api_state.get("running"):
            api_state["analyses_count"] += 1
            api_state["total_logs"] += len(log_files)
            api_state["total_lines"] += total_lines
            api_state["total_occurrences"] += len(all_results)
    except Exception as e:
        status_var.set(f"Error: {e}")
        messagebox.showerror("Error", str(e))
    finally:
        if tmp_xlsx and os.path.exists(tmp_xlsx.name):
            try:
                os.remove(tmp_xlsx.name)
            except Exception:
                pass
        btn_run.config(state="normal")


def browse_xlsx(var):
    path = filedialog.askopenfilename(
        title="Seleccionar base de eventos (.xlsx)",
        filetypes=[("Excel", "*.xlsx")],
    )
    if path:
        var.set(path)


def browse_output(var, exe_dir=None):
    _reports_dir = os.path.join(exe_dir or os.getcwd(), "reportes_html")
    path = filedialog.asksaveasfilename(
        title="Guardar reporte como...",
        initialdir=_reports_dir if os.path.isdir(_reports_dir) else (exe_dir or ""),
        defaultextension=".html",
        filetypes=[("HTML", "*.html"), ("Excel", "*.xlsx")],
    )
    if path:
        var.set(path)


def main():
    from api_tracker import _now_utc, report_session

    root = tk.Tk()
    root.title("Log Analyzer VL550")
    root.resizable(True, True)

    # ── Estado del tracking API ───────────────────────────────────────────────
    api_state = {
        "running": False,
        "start_time": None,
        "analyses_count": 0,        # cuántos análisis se corrieron
        "total_logs": 0,            # cuántos archivos de log procesados
        "total_lines": 0,           # líneas totales procesadas
        "total_occurrences": 0,     # ocurrencias encontradas
    }

    def toggle_api():
        if not api_state["running"]:
            # Iniciar tracking
            api_state["running"] = True
            api_state["start_time"] = _now_utc()
            api_state["analyses_count"] = 0
            api_state["total_logs"] = 0
            api_state["total_lines"] = 0
            api_state["total_occurrences"] = 0
            btn_api.config(text="⏹  Detener API", style="ApiStop.TButton")
            api_status_var.set(f"🟢 API activa desde {api_state['start_time'][:19]}")
            status_var.set("API tracking iniciado.")
        else:
            # Detener y reportar
            end_time = _now_utc()
            details = {
                "analisis_ejecutados": api_state["analyses_count"],
                "logs_procesados": api_state["total_logs"],
                "lineas_analizadas": api_state["total_lines"],
                "ocurrencias_encontradas": api_state["total_occurrences"],
            }
            api_state["running"] = False
            ok, info = report_session(
                api_state["start_time"], end_time,
                status="success",
                records=api_state["total_lines"],
                details=details,
            )
            api_state["start_time"] = None
            btn_api.config(text="▶  Iniciar API", style="ApiStart.TButton")
            api_status_var.set("⚪ API inactiva")
            if ok:
                status_var.set(
                    f"API reportada (HTTP {info}): "
                    f"{details['analisis_ejecutados']} análisis, "
                    f"{details['logs_procesados']} logs, "
                    f"{details['lineas_analizadas']} líneas."
                )
            else:
                status_var.set(f"Error al reportar API: {info}")

    def on_close():
        # Si el tracking está activo al cerrar, reportar automáticamente
        if api_state["running"]:
            end_time = _now_utc()
            details = {
                "analisis_ejecutados": api_state["analyses_count"],
                "logs_procesados": api_state["total_logs"],
                "lineas_analizadas": api_state["total_lines"],
                "ocurrencias_encontradas": api_state["total_occurrences"],
            }
            try:
                report_session(
                    api_state["start_time"], end_time,
                    status="success", records=api_state["total_lines"],
                    details=details,
                )
            except Exception:
                pass
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    pad = {"padx": 10, "pady": 5}

    # ── Variables ─────────────────────────────────────────────────────────────
    output_var = tk.StringVar()
    status_var = tk.StringVar(value="Listo.")

    # Directorio real del exe (o del .py si se corre sin empaquetar)
    if getattr(sys, "frozen", False):
        _exe_dir = os.path.dirname(sys.executable)
    else:
        _exe_dir = os.path.dirname(os.path.abspath(__file__))

    # Buscar xlsx SIEMPRE junto al exe/script, nunca dentro de _MEIPASS
    _default_xlsx = ""
    for _candidate in glob.glob(os.path.join(_exe_dir, "*.xlsx")):
        _default_xlsx = _candidate
        break  # toma el primero que encuentre

    xlsx_var = tk.StringVar(value=_default_xlsx)

    # ── Frame principal ───────────────────────────────────────────────────────
    frm = ttk.Frame(root, padding=16)
    frm.grid(row=0, column=0, sticky="nsew")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    frm.columnconfigure(1, weight=1)

    # Título
    ttk.Label(frm, text="Log Analyzer — VL550 / CGI",
              font=("Segoe UI", 12, "bold")).grid(
        row=0, column=0, columnspan=3, pady=(0, 12), sticky="w"
    )

    # ── Lista de logs ─────────────────────────────────────────────────────────
    ttk.Label(frm, text="Archivos de log:").grid(row=1, column=0, sticky="nw", **pad)

    list_frame = ttk.Frame(frm)
    list_frame.grid(row=1, column=1, sticky="nsew", **pad)
    list_frame.columnconfigure(0, weight=1)
    frm.rowconfigure(1, weight=1)

    log_listbox = tk.Listbox(list_frame, height=15, selectmode=tk.EXTENDED,
                              font=("Consolas", 9), activestyle="dotbox")
    log_listbox.grid(row=0, column=0, sticky="nsew")

    sb = ttk.Scrollbar(list_frame, orient="vertical", command=log_listbox.yview)
    sb.grid(row=0, column=1, sticky="ns")
    log_listbox.configure(yscrollcommand=sb.set)

    # Botones de la lista
    btn_frame = ttk.Frame(frm)
    btn_frame.grid(row=1, column=2, sticky="n", padx=(0, 10), pady=5)

    def add_logs():
        _logs_dir = os.path.join(_exe_dir, "logs")
        paths = filedialog.askopenfilenames(
            title="Agregar archivos de log",
            initialdir=_logs_dir if os.path.isdir(_logs_dir) else _exe_dir,
            filetypes=[("Todos los archivos", "*.*"),
                       ("Texto", "*.txt"), ("Log", "*.log")],
        )
        existing = list(log_listbox.get(0, tk.END))
        for p in paths:
            if p not in existing:
                log_listbox.insert(tk.END, p)

    def remove_selected():
        for i in reversed(log_listbox.curselection()):
            log_listbox.delete(i)

    def clear_list():
        log_listbox.delete(0, tk.END)

    ttk.Button(btn_frame, text="➕ Agregar", width=12, command=add_logs).pack(pady=2)
    ttk.Button(btn_frame, text="➖ Quitar",  width=12, command=remove_selected).pack(pady=2)
    ttk.Button(btn_frame, text="🗑 Limpiar", width=12, command=clear_list).pack(pady=2)

    # ── Base de eventos ───────────────────────────────────────────────────────
    ttk.Label(frm, text="Base de eventos (.xlsx):").grid(row=2, column=0, sticky="w", **pad)
    ttk.Entry(frm, textvariable=xlsx_var, width=80).grid(row=2, column=1, sticky="ew", **pad)
    ttk.Button(frm, text="Examinar…", command=lambda: browse_xlsx(xlsx_var)).grid(
        row=2, column=2, **pad
    )

    # ── Archivo de salida ─────────────────────────────────────────────────────
    ttk.Label(frm, text="Guardar reporte en:").grid(row=3, column=0, sticky="w", **pad)
    ttk.Entry(frm, textvariable=output_var, width=80).grid(row=3, column=1, sticky="ew", **pad)
    ttk.Button(frm, text="Guardar como…", command=lambda: browse_output(output_var, _exe_dir)).grid(
        row=3, column=2, **pad
    )
    ttk.Label(frm, text="(extensión .html o .xlsx determina el formato)",
              foreground="#666", font=("Segoe UI", 8)).grid(
        row=4, column=1, sticky="w", padx=10
    )

    # ── Separador ─────────────────────────────────────────────────────────────
    ttk.Separator(frm, orient="horizontal").grid(
        row=5, column=0, columnspan=3, sticky="ew", pady=10
    )

    # ── Botón API tracking ────────────────────────────────────────────────────
    api_status_var = tk.StringVar(value="⚪ API inactiva")

    style = ttk.Style()
    style.configure("ApiStart.TButton", foreground="green")
    style.configure("ApiStop.TButton", foreground="red")

    api_frame = ttk.Frame(frm)
    api_frame.grid(row=6, column=0, columnspan=3, pady=(0, 4))

    btn_api = ttk.Button(api_frame, text="▶  Iniciar API", width=18,
                         style="ApiStart.TButton", command=toggle_api)
    btn_api.pack(side="left", padx=(0, 10))

    ttk.Label(api_frame, textvariable=api_status_var,
              font=("Segoe UI", 9)).pack(side="left")

    # ── Separador ─────────────────────────────────────────────────────────────
    ttk.Separator(frm, orient="horizontal").grid(
        row=7, column=0, columnspan=3, sticky="ew", pady=6
    )

    # ── Botón Analizar ────────────────────────────────────────────────────────
    btn_run = ttk.Button(frm, text="▶  Analizar")

    def on_run():
        log_files   = list(log_listbox.get(0, tk.END))
        xlsx_file   = xlsx_var.get().strip()
        output_file = output_var.get().strip()

        if not log_files:
            messagebox.showwarning("Falta dato", "Agregá al menos un archivo de log.")
            return
        missing = [f for f in log_files if not os.path.exists(f)]
        if missing:
            messagebox.showerror("Error", "No se encontraron:\n" + "\n".join(missing))
            return
        if not xlsx_file:
            messagebox.showwarning("Falta dato", "Seleccioná la base de eventos (.xlsx).")
            return
        if not os.path.exists(xlsx_file):
            messagebox.showerror("Error", f"No se encontró el Excel:\n{xlsx_file}")
            return
        if not output_file:
            messagebox.showwarning("Falta dato", "Indicá dónde guardar el reporte.")
            return

        btn_run.config(state="disabled")
        status_var.set("Procesando…")
        threading.Thread(
            target=run_analysis,
            args=(log_files, xlsx_file, output_file, status_var, btn_run, root,
                  api_state),
            daemon=True,
        ).start()

    btn_run.config(command=on_run)
    btn_run.grid(row=8, column=0, columnspan=3, pady=(4, 10))

    # ── Barra de estado ───────────────────────────────────────────────────────
    ttk.Separator(frm, orient="horizontal").grid(
        row=9, column=0, columnspan=3, sticky="ew"
    )
    ttk.Label(frm, textvariable=status_var, foreground="#444",
              font=("Segoe UI", 8)).grid(
        row=10, column=0, columnspan=3, sticky="w", padx=10, pady=(4, 2)
    )

    root.mainloop()


if __name__ == "__main__":
    main()
